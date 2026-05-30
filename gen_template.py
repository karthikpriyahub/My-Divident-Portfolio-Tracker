"""
gen_template.py
Generates DivTracker_Template.xlsx using openpyxl.
Usage:
  python3 gen_template.py '<json_stocks>' <output_path>

  json_stocks : JSON array of {stockName, type, sector, symbol}
  output_path : where to write the .xlsx file
"""
import sys, json, io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ── Args ──────────────────────────────────────────────────────────────────────
stocks_json = sys.argv[1] if len(sys.argv) > 1 else "[]"
rows_json   = sys.argv[2] if len(sys.argv) > 2 else "[]"
out_path    = sys.argv[3] if len(sys.argv) > 3 else "-"

stocks    = json.loads(stocks_json)
exist_rows = json.loads(rows_json)

FALLBACK = [
    # Equity
    {"stockName": "Coal India Ltd",           "type": "Equity", "sector": "Energy",         "symbol": "COALINDIA"    },
    {"stockName": "HCL Technologies Ltd",     "type": "Equity", "sector": "IT",             "symbol": "HCLTECH"      },
    {"stockName": "ITC Limited",              "type": "Equity", "sector": "FMCG",           "symbol": "ITC"          },
    {"stockName": "Power Finance Corp",       "type": "Equity", "sector": "Finance",        "symbol": "PFC"          },
    {"stockName": "REC Limited",              "type": "Equity", "sector": "Finance",        "symbol": "RECLTD"       },
    {"stockName": "Castrol India Ltd",        "type": "Equity", "sector": "Energy",         "symbol": "CASTROLIND"   },
    {"stockName": "Vedanta Limited",          "type": "Equity", "sector": "Metals",         "symbol": "VEDL"         },
    # REITs
    {"stockName": "Nexus Select Trust",       "type": "REIT",   "sector": "Real Estate",    "symbol": "NEXUSSELECT"  },
    {"stockName": "Mindspace Business Parks", "type": "REIT",   "sector": "Real Estate",    "symbol": "MINDSPACE"    },
    {"stockName": "Brookfield India REIT",    "type": "REIT",   "sector": "Real Estate",    "symbol": "BIRET"        },
    {"stockName": "Embassy Office Parks",     "type": "REIT",   "sector": "Real Estate",    "symbol": "EMBASSYOFFICE"},
    # InvITs
    {"stockName": "Bharat Highways InvIT",    "type": "InvIT",  "sector": "Infrastructure", "symbol": "BHARATHWY"    },
    {"stockName": "IRB InvIT Fund",           "type": "InvIT",  "sector": "Infrastructure", "symbol": "IRBINVIT"     },
    {"stockName": "India Grid Trust",         "type": "InvIT",  "sector": "Infrastructure", "symbol": "INDIGRID"     },
    {"stockName": "PowerGrid Infra Invest",   "type": "InvIT",  "sector": "Infrastructure", "symbol": "PGCIL"        },
]

master = stocks if stocks else FALLBACK
MAX_ROWS = 500

# ── Styles ────────────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill("solid", fgColor="1F3864")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
AUTO_FILL  = PatternFill("solid", fgColor="E8F4FD")
INPUT_FILL = PatternFill("solid", fgColor="FFFBE6")
CALC_FILL  = PatternFill("solid", fgColor="E8F8EE")
ALT_FILL   = PatternFill("solid", fgColor="F5F9FF")
HDR_ALIGN  = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN       = Side(style="thin", color="CCCCCC")
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADERS = [
    "stockName", "type", "sector", "symbol",
    "qty", "avgPrice", "currentPrice",
    "year", "month",
    "divAmtPerShare", "divQty",   # user inputs
    "grossDiv", "netDiv",         # grossDiv = auto; netDiv = user types
    # tds auto-calculated server-side; notes removed
]
COL_WIDTHS = [26, 12, 14, 13, 8, 12, 14, 7, 7, 18, 10, 14, 14]
COL = {h: i+1 for i, h in enumerate(HEADERS)}

DEC3 = '#,##0.000'   # 3-decimal number format for dividend columns

# ── Workbook ──────────────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: Tracker (active) ─────────────────────────────────────────────────
ws = wb.active
ws.title = "Tracker"
ws.freeze_panes = "A2"
ws.sheet_properties.tabColor = "1F3864"

# header row
for col_idx, (hdr, w) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
    cell = ws.cell(row=1, column=col_idx, value=hdr)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = HDR_ALIGN
    cell.border    = BORDER
    ws.column_dimensions[get_column_letter(col_idx)].width = w

ws.row_dimensions[1].height = 22

# sample data rows (use existing rows if passed, else use defaults)
SAMPLES_DEFAULT = [
    # stockName, type, sector, symbol, qty, avgPrice, currentPrice, year, month, divAmtPerShare, divQty, grossDiv, netDiv
    ["Coal India Ltd",     "Equity","Energy",        "COALINDIA",   100, 450,  480,  2025, 1, 5.000,  100, 500.000, 450.000],
    ["ITC Limited",        "Equity","FMCG",          "ITC",         200, 400,  430,  2025, 2, 7.500,  200,1500.000,1350.000],
    ["Nexus Select Trust", "REIT",  "Real Estate",   "NEXUSSELECT", 150, 130,  140,  2025, 3, 5.250,  150, 787.500, 787.500],
    ["India Grid Trust",   "InvIT", "Infrastructure","INDIGRID",    300, 135,  145,  2025, 3, 3.500,  300,1050.000,1050.000],
]

if exist_rows:
    SAMPLES = [
        [r.get("stockName",""), r.get("type",""), r.get("sector",""), r.get("symbol",""),
         r.get("qty",0), r.get("avgPrice",0), r.get("currentPrice",0),
         r.get("year",0), r.get("month",0),
         r.get("divAmtPerShare",0), r.get("divQty",0),
         r.get("grossDiv",0), r.get("netDiv",0)]
        for r in exist_rows
    ]
else:
    SAMPLES = SAMPLES_DEFAULT

for row_idx, row_data in enumerate(SAMPLES, start=2):
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border    = BORDER
        cell.alignment = Alignment(vertical="center")
        if col_idx in (COL["type"], COL["sector"], COL["symbol"]):
            cell.fill = AUTO_FILL
        elif col_idx in (COL["divAmtPerShare"], COL["divQty"], COL["netDiv"]):
            cell.fill = INPUT_FILL
        elif col_idx == COL["grossDiv"]:
            cell.fill = CALC_FILL
        # 3-decimal format for dividend money columns
        if col_idx in (COL["divAmtPerShare"], COL["grossDiv"], COL["netDiv"]):
            cell.number_format = DEC3

# ── Formulas for blank rows ────────────────────────────────────────
# Columns: B=type(VLOOKUP), C=sector(VLOOKUP), D=symbol(VLOOKUP)
#          J=divAmtPerShare(user), K=divQty(user)
#          L=grossDiv(=J*K auto), M=netDiv(user)
# TDS is NOT in the template — server auto-calculates as grossDiv-netDiv
n_stocks = len(master)
sm_range = f"StockMaster!$A$2:$D${n_stocks + 1}"
data_end = len(SAMPLES) + 2

for row in range(data_end, MAX_ROWS + 1):
    a = f"A{row}"

    # VLOOKUP auto-fills (light blue)
    for col_letter, col_idx, lookup_col in (
        ("B", COL["type"],   2),
        ("C", COL["sector"], 3),
        ("D", COL["symbol"], 4),
    ):
        cell = ws.cell(row=row, column=col_idx,
                       value=f'=IFERROR(VLOOKUP({a},{sm_range},{lookup_col},0),"")')
        cell.fill = AUTO_FILL; cell.border = BORDER

    # divAmtPerShare (J) — yellow user input, 3-decimal format
    cj = ws.cell(row=row, column=COL["divAmtPerShare"])
    cj.fill = INPUT_FILL; cj.border = BORDER; cj.number_format = DEC3

    # divQty (K) — yellow user input
    ck = ws.cell(row=row, column=COL["divQty"])
    ck.fill = INPUT_FILL; ck.border = BORDER

    # grossDiv (L) = divAmtPerShare × divQty — green auto, 3-decimal
    j, k = f"J{row}", f"K{row}"
    cl = ws.cell(row=row, column=COL["grossDiv"],
                 value=f"=IF(OR({j}<>0,{k}<>0),{j}*{k},0)")
    cl.fill = CALC_FILL; cl.border = BORDER; cl.number_format = DEC3

    # netDiv (M) — yellow user input, 3-decimal
    cm = ws.cell(row=row, column=COL["netDiv"])
    cm.fill = INPUT_FILL; cm.border = BORDER; cm.number_format = DEC3

# ── Dropdown data validation on column A (stockName) ─────────────────────────
dv = DataValidation(
    type="list",
    formula1=f"StockMaster!$A$2:$A${n_stocks + 1}",
    allow_blank=True,
    showDropDown=False,   # False = show the arrow
    showErrorMessage=True,
    errorTitle="Invalid stock",
    error="Please select a stock from the dropdown list.",
)
dv.sqref = f"A2:A{MAX_ROWS}"
ws.add_data_validation(dv)

# ── Sheet 2: StockMaster (hidden — drives dropdown + VLOOKUP) ────────────────
sm = wb.create_sheet("StockMaster")
sm_headers = ["stockName", "type", "sector", "symbol"]
sm_widths  = [26, 13, 14, 13]

for col_idx, (hdr, w) in enumerate(zip(sm_headers, sm_widths), start=1):
    cell = sm.cell(row=1, column=col_idx, value=hdr)
    cell.font      = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    sm.column_dimensions[get_column_letter(col_idx)].width = w

for row_idx, s in enumerate(master, start=2):
    sm.cell(row=row_idx, column=1, value=s.get("stockName", ""))
    sm.cell(row=row_idx, column=2, value=s.get("type",      "Equity"))
    sm.cell(row=row_idx, column=3, value=s.get("sector",    ""))
    sm.cell(row=row_idx, column=4, value=s.get("symbol",    ""))

# Hide StockMaster from the tab bar
sm.sheet_state = "hidden"

# ── Write output ──────────────────────────────────────────────────────────────
if out_path == "-":
    buf = io.BytesIO()
    wb.save(buf)
    sys.stdout.buffer.write(buf.getvalue())
else:
    wb.save(out_path)
